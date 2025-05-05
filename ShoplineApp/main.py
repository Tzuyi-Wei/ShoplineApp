import pandas as pd
import datetime
import os
import openpyxl
import re
import tkinter as tk
from tkinter import filedialog, messagebox

translation_map = {
    "Taiwan": "台灣",
    "New Taipei City": "新北市",
    "Taipei City": "台北市",
    "Taoyuan City": "桃園市",
    "Taichung City": "台中市",
    "Tainan City": "台南市",
    "Kaohsiung City": "高雄市",
}

def process_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        de_col_idx = 109
        sheet.insert_cols(de_col_idx)
        sheet.cell(row=1, column=de_col_idx, value="折扣總金額")
        for row in range(2, sheet.max_row + 1):
            formula = "=SUM(CZ{}:DD{})".format(row, row)
            sheet.cell(row=row, column=de_col_idx, value=formula)
        bh_col_idx = 60
        for row in range(2, sheet.max_row + 1):
            full_address = sheet.cell(row=row, column=bh_col_idx).value
            if full_address and full_address.startswith("台灣 "):
                updated_address = re.sub(r"^台灣 \d{3,5} ", "", full_address)
                if updated_address.strip() != "台灣":
                    sheet.cell(row=row, column=bh_col_idx, value=updated_address)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=55, max_col=60):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for eng, zh in translation_map.items():
                        if eng in cell.value:
                            cell.value = cell.value.replace(eng, zh)
        today_date = datetime.datetime.now().strftime("%m%d")
        folder_path = os.path.dirname(file_path)
        output_filename = f"{today_date}_Shopline訂單.xlsx"
        output_path = os.path.join(folder_path, output_filename)
        wb.save(output_path)
        messagebox.showinfo("完成", f"處理完成！已儲存：\n{output_path}")
    except Exception as e:
        messagebox.showerror("錯誤", f"處理失敗：\n{str(e)}")

root = tk.Tk()
root.withdraw()
process_excel()
# Trigger GitHub Actions
